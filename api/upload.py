"""
upload.py — /api/upload handler.

Accepts multipart POST with fields:
  company_name, sector, period, report_type, file_type, file, extracted_text

For Excel: extracts text via openpyxl.
For PDF: uses extracted_text provided by the client (pdf.js).

Calls Claude to extract structured financials, then upserts into Supabase:
  companies, reports, financials tables.
"""

import io
import json
import os
import re
import traceback
from datetime import date


COMPANY_RE = re.compile(r"^[A-Za-z0-9\s\-\.&,()]{1,80}$")
MAX_CONTENT_CHARS = 80_000


def handle_upload():
    try:
        return _handle_upload_inner()
    except Exception as e:
        print(f"[upload] Unhandled exception:\n{traceback.format_exc()}")
        return _json({"success": False, "error": str(e)}, 500)


def _handle_upload_inner():
    from flask import request

    company_name   = (request.form.get("company_name") or "").strip()
    sector         = (request.form.get("sector") or "").strip()
    period         = (request.form.get("period") or "").strip()
    report_type    = (request.form.get("report_type") or "").strip().lower()
    file_type      = (request.form.get("file_type") or "").strip().lower()
    extracted_text = (request.form.get("extracted_text") or "").strip()
    uploaded_file  = request.files.get("file")

    print(f"[upload] company={company_name!r} sector={sector!r} period={period!r} "
          f"report_type={report_type!r} file_type={file_type!r}")

    # --- Validation ---
    if not company_name or not COMPANY_RE.match(company_name):
        return _json({"success": False, "error": "Invalid or missing company_name"}, 400)
    if not sector:
        return _json({"success": False, "error": "sector is required"}, 400)
    if not period:
        return _json({"success": False, "error": "period is required (e.g. '2024_Q1' or '2024_annual')"}, 400)
    if report_type not in ("quarterly", "annual"):
        return _json({"success": False, "error": "report_type must be 'quarterly' or 'annual'"}, 400)
    if file_type not in ("excel", "pdf"):
        return _json({"success": False, "error": "file_type must be 'excel' or 'pdf'"}, 400)
    if uploaded_file is None:
        return _json({"success": False, "error": "file is required"}, 400)

    # --- Extract content ---
    try:
        if file_type == "excel":
            content = _extract_excel(uploaded_file)
        else:
            content = extracted_text
    except Exception as e:
        return _json({"success": False, "error": f"Failed to read file: {e}"}, 422)

    if not content or len(content.strip()) < 100:
        return _json({"success": False, "error": "Extracted content too short — check the file"}, 422)

    # --- Claude extraction ---
    try:
        financial_data = _extract_with_claude(content)
    except Exception as e:
        print(f"[upload] Claude error:\n{traceback.format_exc()}")
        return _json({"success": False, "error": f"AI extraction failed: {e}"}, 500)

    # --- Supabase inserts ---
    try:
        period_date = _period_to_date(period)
        filename = uploaded_file.filename or f"{company_name}_{period}.{file_type}"
        _save_to_supabase(company_name, sector, period, report_type, file_type,
                          filename, financial_data, period_date)
    except Exception as e:
        print(f"[upload] Supabase error:\n{traceback.format_exc()}")
        return _json({"success": False, "error": f"Database error: {e}"}, 500)

    return _json({"success": True, "message": "Data uploaded successfully"})


# ---------------------------------------------------------------------------
# Excel extraction
# ---------------------------------------------------------------------------

def _extract_excel(file_storage) -> str:
    import openpyxl
    data = file_storage.read()
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    lines = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        lines.append(f"\n=== Sheet: {sheet_name} ===")
        for row in ws.iter_rows(values_only=True):
            if any(cell is not None for cell in row):
                lines.append("\t".join("" if c is None else str(c) for c in row))
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Claude API extraction
# ---------------------------------------------------------------------------

_CLAUDE_PROMPT = """You are a financial data extraction assistant for Nordic bank reports.
Extract data from the report below and return ONLY a valid JSON object with exactly these keys.
Map any field names in the report to the closest matching key in the schema.
Any field not found must be null. Numbers should be plain numbers (no commas, no units).
Return ONLY the JSON object — no markdown, no explanation.

Schema:
{
  "income_statement": {
    "net_interest_income": null,
    "net_commission_income": null,
    "net_gains_on_financial_items": null,
    "other_income": null,
    "total_income": null,
    "staff_costs": null,
    "other_expenses": null,
    "total_expenses": null,
    "profit_before_impairments": null,
    "credit_impairments": null,
    "profit_before_tax": null,
    "tax": null,
    "net_profit": null
  },
  "balance_sheet": {
    "loans_to_customers": null,
    "deposits_from_customers": null,
    "total_assets": null,
    "total_equity": null,
    "risk_exposure_amount": null
  },
  "key_ratios": {
    "return_on_equity_pct": null,
    "cost_income_ratio": null,
    "eps_diluted": null,
    "cet1_capital_ratio_pct": null,
    "credit_impairment_ratio_pct": null,
    "net_interest_margin_pct": null
  },
  "currency": null
}
"""


def _extract_with_claude(content: str) -> dict:
    import anthropic
    client = anthropic.Anthropic(api_key=os.environ["ANTHROPIC_API_KEY"])

    truncated = content[:MAX_CONTENT_CHARS]
    message = client.messages.create(
        model="claude-sonnet-4-20250514",
        max_tokens=2048,
        messages=[
            {"role": "user", "content": f"{_CLAUDE_PROMPT}\n\nReport content:\n{truncated}"}
        ],
    )

    raw = message.content[0].text.strip()
    # Strip markdown fences if present
    raw = re.sub(r"^```(?:json)?\s*", "", raw)
    raw = re.sub(r"\s*```$", "", raw)

    return json.loads(raw)


# ---------------------------------------------------------------------------
# Period → date
# ---------------------------------------------------------------------------

_QUARTER_END = {"q1": (3, 31), "q2": (6, 30), "q3": (9, 30), "q4": (12, 31)}


def _period_to_date(period: str) -> str:
    parts = period.lower().split("_")
    try:
        year = int(parts[0])
    except (ValueError, IndexError):
        raise ValueError(f"Cannot parse year from period {period!r}")

    suffix = parts[1] if len(parts) > 1 else "annual"
    month, day = _QUARTER_END.get(suffix, (12, 31))
    return date(year, month, day).isoformat()


# ---------------------------------------------------------------------------
# Supabase persistence
# ---------------------------------------------------------------------------

def _save_to_supabase(company_name, sector, period, report_type, file_type,
                      filename, financial_data, period_date):
    from supabase import create_client

    url = os.environ["SUPABASE_URL"]
    key = os.environ["SUPABASE_ANON_KEY"]
    sb = create_client(url, key)

    # Upsert company (ignore if already exists)
    sb.table("companies").upsert(
        {"company_name": company_name, "sector": sector},
        on_conflict="company_name",
        ignore_duplicates=True,
    ).execute()

    # Insert report record
    sb.table("reports").insert({
        "company_name": company_name,
        "period":       period,
        "report_type":  report_type,
        "file_type":    file_type,
        "filename":     filename,
    }).execute()

    # Insert extracted financials
    sb.table("financials").insert({
        "company_name":     company_name,
        "period":           period,
        "report_type":      report_type,
        "period_date":      period_date,
        "income_statement": financial_data.get("income_statement"),
        "balance_sheet":    financial_data.get("balance_sheet"),
        "key_ratios":       financial_data.get("key_ratios"),
        "currency":         financial_data.get("currency"),
    }).execute()


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _json(body: dict, status: int = 200):
    from flask import jsonify
    return jsonify(body), status
